#!/usr/bin/env python3
# ╔══════════════════════════════════════════════════════════════════════════╗
# ║  JasperBot — GitHub Actions                                             ║
# ║  1 driver → 1 login → 4 tab (Cell 2/3/4/5)                             ║
# ║  Rebuilt dari notebook jasper_lagi.ipynb (versi Colab yang berhasil)    ║
# ╚══════════════════════════════════════════════════════════════════════════╝

import os, sys, time, glob, shutil, re, json, traceback
from datetime import datetime, timezone, timedelta

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC



import gspread
from google.oauth2.service_account import Credentials
import openpyxl, xlrd

# =============================================================================
# KONFIGURASI
# =============================================================================
USERNAME     = "muhammad.prasetyo"
PASSWORD     = "Adminhqacc12"
BASE_URL     = "http://report.tangki.id/jasperserver"
GSHEET_ID    = "1BTAVmWs-9GZpJcO2Kv2zEtV2jy680wHASboIeArqb9U"
WIB          = timezone(timedelta(hours=7))
TODAY_STR    = datetime.now(WIB).strftime("%Y-%m-%d")
TODAY_LABEL  = datetime.now(WIB).strftime("%Y%m%d")

DOWNLOAD_DIR    = "/tmp/jasper_downloads/"
C5_DOWNLOAD_DIR = "/tmp/erp_downloads/"
FOLDER_OUT      = "/tmp/jasper_exports/"
SEARCH_DIRS     = [DOWNLOAD_DIR, "/tmp", os.path.expanduser("~")]
C5_SEARCH_DIRS  = [C5_DOWNLOAD_DIR, "/tmp"]
EXTENSIONS      = ["*.xls", "*.xlsx", "*.XLS", "*.XLSX"]

ERP_URL  = "https://erp.tangki.id/webui/index.zul"
ERP_USER = "muhammad.prasetyo"
ERP_PASS = "Adminhqacc12"

for d in [DOWNLOAD_DIR, C5_DOWNLOAD_DIR, FOLDER_OUT]:
    os.makedirs(d, exist_ok=True)

# =============================================================================
# GOOGLE SHEETS AUTH  (service account dari env GSHEET_CREDENTIALS_B64)
# =============================================================================
def init_gc():
    import base64
    b64 = os.environ.get("GSHEET_CREDENTIALS_B64", "")
    if not b64:
        raise RuntimeError("GSHEET_CREDENTIALS_B64 tidak ditemukan di environment!")
    creds = Credentials.from_service_account_info(
        json.loads(base64.b64decode(b64).decode()),
        scopes=["https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"])
    return gspread.authorize(creds)

# =============================================================================
# DRIVER
# =============================================================================
def make_driver(download_dir=DOWNLOAD_DIR):
    opts = webdriver.ChromeOptions()
    opts.add_argument("--headless")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--ignore-certificate-errors")
    opts.add_argument("--ignore-ssl-errors")
    opts.accept_insecure_certs = True
    opts.add_experimental_option("prefs", {
        "download.default_directory":   download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade":   True,
        "safebrowsing.enabled":         True,
        "profile.default_content_settings.popups": 0,
        "profile.content_settings.exceptions.automatic_downloads.*.setting": 1,
    })
    d = webdriver.Chrome(options=opts)
    # Beri halaman waktu cukup untuk load (fix ReadTimeoutError)
    d.set_script_timeout(120)
    d.implicitly_wait(0)           # jangan implicit wait, pakai explicit
    return d

# =============================================================================
# HELPER FUNCTIONS  (identik dengan notebook)
# =============================================================================
def clean_value(val):
    if val is None: return ''
    if isinstance(val, (int, float)):
        try: return int(val) if val == int(val) else val
        except: return val
    s = str(val).strip()
    if not s: return ''
    try:
        c = s
        if re.match(r'^-?\d{1,3}(,\d{3})+(\.\d+)?$', c): c = c.replace(',', '')
        elif re.match(r'^-?\d{1,3}(\.\d{3})+(,\d+)?$', c):
            c = c.replace('.', '').replace(',', '.')
        f = float(c)
        return int(f) if f == int(f) else f
    except: return s

def wait_ready(driver, t=30):
    WebDriverWait(driver, t).until(
        lambda d: d.execute_script("return document.readyState") == "complete")
    time.sleep(2)

def do_click(driver, el, x, y):
    driver.execute_script(
        "var el=arguments[0],x=arguments[1],y=arguments[2];"
        "var o={bubbles:true,cancelable:true,clientX:x,clientY:y,screenX:x,screenY:y,view:window};"
        "['mouseover','mouseenter','mousemove','mousedown','mouseup','click'].forEach("
        "    function(ev){el.dispatchEvent(new MouseEvent(ev,o));});",
        el, x, y)

def trigger_events(driver, inp):
    driver.execute_script(
        "var el=arguments[0];"
        "['focus','input','change'].forEach(function(ev){"
        "    el.dispatchEvent(new Event(ev,{bubbles:true}));});"
        "el.dispatchEvent(new KeyboardEvent('keyup',{bubbles:true,key:'Tab',keyCode:9}));"
        "el.dispatchEvent(new Event('blur',{bubbles:true}));",
        inp)

def is_loading_visible(driver):
    return driver.execute_script(
        "var els=document.querySelectorAll('*');"
        "for(var i=0;i<els.length;i++){"
        "  var el=els[i]; if(!el.offsetParent) continue;"
        "  if(el.getBoundingClientRect().width<10) continue;"
        "  var t=el.textContent.trim();"
        "  if(t==='Loading...'||(t.startsWith('Loading')&&t.length<30)) return true;"
        "} return false;")

def scan_downloads(search_dirs=None):
    dirs = search_dirs or SEARCH_DIRS
    found = []
    for d in dirs:
        if not os.path.exists(d): continue
        for ext in EXTENSIONS: found += glob.glob(os.path.join(d, ext))
    now = time.time()
    return [f for f in found if now - os.path.getmtime(f) < 300]

def do_login(driver):
    print("  → Login ...")
    driver.get(f"{BASE_URL}/login.html")
    wait_ready(driver)
    driver.find_element(By.ID, "j_username").send_keys(USERNAME)
    try:
        p = driver.find_element(By.ID, "j_password_pseudo")
        p.click()
        time.sleep(0.5)
    except:
        p = driver.find_element(By.ID, "j_password")
    p.send_keys(PASSWORD)
    try:
        WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.ID, "submitButton"))).click()
    except TimeoutException:
        p.send_keys(Keys.RETURN)
    time.sleep(5)
    print("  ✅ Login OK")

def click_apply_dialog(driver):
    print("\n  🔵 Klik Apply ...")
    btn = None
    for sel in [(By.ID, "apply"), (By.CSS_SELECTOR, "button#apply"),
                (By.XPATH, "//button[normalize-space()='Apply']"),
                (By.XPATH, "//input[@value='Apply']")]:
        try:
            el = driver.find_element(*sel)
            if el.is_displayed(): btn = el; break
        except: continue
    if not btn: print("  ❌ Apply tidak ditemukan!"); return False
    driver.execute_script("arguments[0].scrollIntoView({block:'nearest'});", btn); time.sleep(0.5)
    br = driver.execute_script(
        "var r=arguments[0].getBoundingClientRect();"
        "return {x:Math.round(r.x+r.width/2),y:Math.round(r.y+r.height/2)};", btn)
    bx, by = br['x'], br['y']
    driver.execute_script("arguments[0].click();", btn); time.sleep(2)
    if is_loading_visible(driver): print("  ✅ Apply S1!"); return True
    do_click(driver, btn, bx, by); time.sleep(2)
    if is_loading_visible(driver): print("  ✅ Apply S2!"); return True
    ActionChains(driver).move_to_element(btn).pause(0.5).click().perform(); time.sleep(2)
    if is_loading_visible(driver): print("  ✅ Apply S3!"); return True
    print("  ⚠️ Loading tidak terdeteksi — lanjut ..."); return True

def wait_loading(driver):
    print("\n  ⏳ Tunggu loading muncul (max 60s) ...")
    appeared = False
    for i in range(60):
        time.sleep(1)
        if is_loading_visible(driver):
            print(f"  ✅ Loading muncul [{i+1}s]"); appeared = True; break
        if (i+1) % 10 == 0: print(f"    [{i+1}s] menunggu ...")
    if not appeared: print("  ⚠️ Loading tidak muncul — lanjut ...")
    print("  ⏳ Tunggu loading selesai ...")
    tick = 0
    while True:
        time.sleep(5); tick += 1
        if not is_loading_visible(driver):
            print(f"  ✅ Loading selesai ~{tick*5}s"); break
        if tick % 12 == 0: print(f"    [{tick*5}s] masih loading ...")
        if tick > 120: print("  ⚠️ Timeout 600s — lanjut ..."); break

def export_xlsx(driver, search_dirs=None):
    print("\n  📤 Export XLSX ...")
    driver.switch_to.default_content(); time.sleep(2)

    def dropdown_open():
        return driver.execute_script(
            "var items=document.querySelectorAll('a,li');"
            "for(var i=0;i<items.length;i++){"
            "  var t=items[i].textContent.trim();"
            "  if(t==='XLSX'||t==='Excel'){"
            "    var b=items[i].getBoundingClientRect();"
            "    if(b.width>20&&b.height>5) return true;}}"
            "return false;")

    ex, ey = 137, 96; export_el = None
    for sel in ["button[title='Export']", "a[title='Export']", "li[title='Export']",
                ".jr-mButton-export"]:
        try:
            el = driver.find_element(By.CSS_SELECTOR, sel)
            if el.is_displayed():
                b = driver.execute_script(
                    "var r=arguments[0].getBoundingClientRect();"
                    "return {x:Math.round(r.x+r.width/2),y:Math.round(r.y+r.height/2)};", el)
                export_el = el; ex, ey = b['x'], b['y']
                print(f"  → Export btn: ({ex},{ey})"); break
        except: continue
    if not export_el:
        export_el = driver.execute_script(f"return document.elementFromPoint({ex},{ey});")

    for mname, fn in [
        ("ActionChains", lambda e: ActionChains(driver).move_to_element(e).pause(0.5).click().perform()),
        ("JS click",     lambda e: driver.execute_script("arguments[0].click();", e)),
    ]:
        try: fn(driver.execute_script(f"return document.elementFromPoint({ex},{ey});") or export_el)
        except Exception as err: print(f"    ⚠️ {mname}: {err}")
        time.sleep(3)
        if dropdown_open(): print(f"  ✅ Dropdown [{mname}]"); break

    if not dropdown_open():
        for hx, hy in [(120, 96), (137, 78), (120, 78), (124, 112)]:
            el = driver.execute_script(f"return document.elementFromPoint({hx},{hy});")
            if not el: continue
            do_click(driver, el, hx, hy); time.sleep(3)
            if dropdown_open(): ex, ey = hx, hy; break

    if not dropdown_open(): print("  ❌ Dropdown tidak bisa dibuka!"); return None

    raw = driver.execute_script(
        "var res=[];"
        "document.querySelectorAll('a,li,span,div,button').forEach(function(el){"
        "  if(!el.offsetParent) return;"
        "  var b=el.getBoundingClientRect();"
        "  if(b.width<20||b.height<5) return;"
        "  var t=el.textContent.trim();"
        "  if(t.length>0&&t.length<60)"
        "    res.push({text:t,x:Math.round(b.x+b.width/2),y:Math.round(b.y+b.height/2)});"
        "}); return res;")
    items = [it for it in raw if 150 < it['y'] < 700 and it['x'] < 350]
    xlsx = None
    for prio in ['XLSX', 'Excel']:
        for it in items:
            if it['text'].strip() == prio: xlsx = it; print(f"  ✅ '{prio}' ({it['x']},{it['y']})"); break
        if xlsx: break
    if not xlsx:
        for it in items:
            if it['text'].lower().startswith('xlsx') and 'paginated' not in it['text'].lower():
                xlsx = it; break
    if not xlsx: print("  ❌ XLSX tidak ditemukan!"); return None

    ix, iy = xlsx['x'], xlsx['y']
    driver.execute_script(
        "var x=arguments[0],y=arguments[1],el=document.elementFromPoint(x,y); if(!el) return;"
        "var o={bubbles:true,cancelable:true,view:window,clientX:x,clientY:y,"
        "       screenX:x,screenY:y,button:0,buttons:0};"
        "el.dispatchEvent(new MouseEvent('mouseover',o));"
        "el.dispatchEvent(new MouseEvent('mouseenter',o));"
        "el.dispatchEvent(new MouseEvent('mousemove',o));", ix, iy); time.sleep(0.8)
    driver.execute_script(
        "var x=arguments[0],y=arguments[1],el=document.elementFromPoint(x,y); if(!el) return;"
        "var o={bubbles:true,cancelable:true,view:window,clientX:x,clientY:y,"
        "       screenX:x,screenY:y,button:0,buttons:1};"
        "el.dispatchEvent(new MouseEvent('mousedown',o));"
        "el.dispatchEvent(new MouseEvent('mouseup',o));"
        "el.dispatchEvent(new MouseEvent('click',o));", ix, iy); time.sleep(2)

    if not dropdown_open(): print("  ✅ XLSX diklik [A]")
    else:
        el = driver.execute_script(f"return document.elementFromPoint({ix},{iy});")
        if el: ActionChains(driver).move_to_element(el).pause(1.0).click().perform()
        time.sleep(2)
        if not dropdown_open(): print("  ✅ XLSX diklik [B]")
        else:
            el = driver.execute_script(f"return document.elementFromPoint({ix},{iy});")
            if el: driver.execute_script("arguments[0].click();", el); time.sleep(2)

    print("  ⏳ Menunggu file download (max 120s) ...")
    sdirs = search_dirs or SEARCH_DIRS
    for i in range(24):
        time.sleep(5); fresh = scan_downloads(sdirs)
        if fresh:
            f = max(fresh, key=os.path.getmtime)
            print(f"  ✅ Download: {f}  ({os.path.getsize(f):,} bytes)"); return f
        if (i+1) % 6 == 0: print(f"    [{(i+1)*5}s] belum ada file ...")
        else: print(f"    [{(i+1)*5}s] menunggu ...")
    print("  ❌ Timeout download!"); return None

def save_to_export(local_file, name_prefix):
    """Simpan backup ke /tmp/jasper_exports/ (ganti save_to_drive untuk Actions)."""
    if not local_file or not os.path.exists(local_file): return None
    ext  = os.path.splitext(local_file)[1]
    dest = os.path.join(FOLDER_OUT, f"{name_prefix}_{TODAY_LABEL}{ext}")
    if os.path.exists(dest): os.remove(dest)
    shutil.copy2(local_file, dest); print(f"  ✅ Export backup: {dest}"); return dest

def save_to_gsheet(gc, local_file, tab, label):
    if not local_file or not os.path.exists(local_file): return None
    try:
        wb   = openpyxl.load_workbook(local_file, data_only=True)
        data = [[clean_value(c) for c in row] for row in wb.active.iter_rows(values_only=True)]
        total_rows, total_cols = len(data), len(data[0]) if data else 0
        print(f"  → {total_rows} baris × {total_cols} kolom")
        sh = gc.open_by_key(GSHEET_ID)
        try: wsg = sh.worksheet(tab)
        except gspread.exceptions.WorksheetNotFound:
            wsg = sh.add_worksheet(title=tab, rows=100, cols=26)
            print(f"  → Tab '{tab}' dibuat")
        wsg.resize(rows=total_rows+10, cols=total_cols+5); time.sleep(1)
        wsg.clear(); time.sleep(0.5)
        for start in range(0, total_rows, 500):
            batch = data[start:start+500]
            wsg.update(range_name=f"A{start+1}", values=batch)
            print(f"  → Upload {start+1}–{start+len(batch)}"); time.sleep(1)
        try: info_ws = sh.worksheet("Info")
        except gspread.exceptions.WorksheetNotFound:
            info_ws = sh.add_worksheet(title="Info", rows=20, cols=5)
        info_ws.update(range_name="A1", values=[
            [f"Last Updated ({tab})", datetime.now(WIB).strftime("%Y-%m-%d %H:%M:%S WIB")],
            [f"Total Rows ({tab})",   total_rows],
            [f"Total Cols ({tab})",   total_cols],
            ["Source", f"{label} {TODAY_STR}"],
        ])
        url = f"https://docs.google.com/spreadsheets/d/{GSHEET_ID}"
        print(f"  ✅ GSheet tab '{tab}': {url}")
        try:
            if local_file and os.path.exists(local_file):
                os.remove(local_file)
                print(f"  🗑️  File lokal dihapus: {os.path.basename(local_file)}")
        except Exception as ce:
            print(f"  ⚠️  Gagal hapus lokal: {ce}")
        return url
    except Exception as e:
        print(f"  ❌ GSheet error: {e}\n{traceback.format_exc()}"); return None

def bot_footer(export_path, gsheet_url, tab):
    print(f"\n{'='*60}\n  🎉 SELESAI!")
    if export_path:  print(f"  📁 Backup : {export_path}")
    if gsheet_url:   print(f"  📊 GSheet : {gsheet_url}  (tab: {tab})")
    print(f"{'='*60}")

def open_new_tab(driver):
    driver.execute_script("window.open('about:blank', '_blank');")
    driver.switch_to.window(driver.window_handles[-1])

# =============================================================================
# CELL 2 — Material Transaction Summary → tab "Data"
# =============================================================================
BOT74_REPORT_URL = (
    f"{BASE_URL}/flow.html?_flowId=viewReportFlow"
    "&reportUnit=/iDempiere/Inventory/Stock/MaterialTransactionSummary"
    "&standAlone=true"
)
BOT74_WAREHOUSE_GROUP = "SCM WHS POK"

def fill_date_v74(driver, label, index):
    print(f"  📅  {label} → '{TODAY_STR}'")
    driver.switch_to.default_content()
    try: driver.execute_script(
        "var dp=document.querySelector('.ui-datepicker');if(dp)dp.style.display='none';")
    except: pass
    time.sleep(0.3)
    inps = driver.find_elements(By.CSS_SELECTOR, "input.date.hasDatepicker")
    if index >= len(inps): print(f"  ❌  index {index} tidak ada!"); return False
    inp = inps[index]
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp); time.sleep(0.4)
    try:
        ActionChains(driver).move_to_element(inp).click().perform(); time.sleep(0.3)
        inp.send_keys(Keys.CONTROL+"a"); time.sleep(0.1)
        inp.send_keys(Keys.DELETE);     time.sleep(0.1)
        inp.send_keys(TODAY_STR);       time.sleep(0.3)
        inp.send_keys(Keys.TAB);        time.sleep(0.5)
        val = inp.get_attribute('value')
        if val and val.strip(): trigger_events(driver, inp); print(f"  ✅  '{val}'"); return True
    except Exception as e: print(f"  ⚠️  S1: {e}")
    try:
        driver.execute_script("""
            var el=arguments[0],v=arguments[1];
            var s=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
            s.call(el,v); el.value=v;
            ['focus','input','change','blur'].forEach(function(e){
                el.dispatchEvent(new Event(e,{bubbles:true}));});
            el.dispatchEvent(new KeyboardEvent('keyup',{bubbles:true,key:'Tab',keyCode:9}));
        """, inp, TODAY_STR); time.sleep(0.5)
        val = inp.get_attribute('value')
        if val and val.strip(): print(f"  ✅  JS '{val}'"); return True
    except Exception as e: print(f"  ⚠️  S2: {e}")
    print(f"  ❌  {label} GAGAL!"); return False

def select_warehouse_group_v74(driver, item_text):
    print(f"  📦  Warehouse Group: '{item_text}'")
    driver.switch_to.default_content()
    try:
        # 1. Scroll ke elemen dropdown
        wg = driver.find_element(By.ID, "WarehouseGroup")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", wg)
        time.sleep(0.8)

        # 2. Klik toggle untuk buka dropdown
        toggle = driver.find_element(By.CSS_SELECTOR, "#WarehouseGroup a.jr-mSingleselect-input")
        tr = driver.execute_script("""
            var r=arguments[0].getBoundingClientRect();
            return {x:Math.round(r.left+r.width/2),y:Math.round(r.top+r.height/2)};
        """, toggle)
        do_click(driver, toggle, tr['x'], tr['y'])
        time.sleep(2.5)

        # 3. Coba klik item (max 3 percobaan)
        for attempt in range(3):
            print(f"    attempt {attempt+1}: mencari '{item_text}'...")

            # Cari elemen dengan text exact match
            match_el = driver.execute_script("""
                var txt = arguments[0];
                var found = null;
                var selectors = [
                    '.jr-mSingleselect-list li a',
                    '.jr-mSingleselect-list li span',
                    '.jr-mSingleselect-list li',
                    'ul li a', 'ul li span', 'ul li'
                ];
                for (var s=0; s<selectors.length; s++) {
                    var els = document.querySelectorAll(selectors[s]);
                    for (var i=0; i<els.length; i++) {
                        var el = els[i];
                        if (el.textContent.trim() !== txt) continue;
                        var r = el.getBoundingClientRect();
                        if (r.width > 0 && r.height > 0 && r.top > 0 && r.top < 1080) {
                            found = el;
                            break;
                        }
                    }
                    if (found) break;
                }
                return found;
            """, item_text)

            if not match_el:
                print(f"    ⚠️  Elemen tidak ditemukan, tunggu...")
                time.sleep(1)
                continue

            print(f"    ✔️  Elemen ditemukan, mencoba klik...")

            # Dispatch full mouse event sequence (paling kompatibel dengan semua framework)
            driver.execute_script("""
                var el = arguments[0];
                el.scrollIntoView({block:'center'});
                ['mouseover','mouseenter','mousemove','mousedown','mouseup','click'].forEach(function(evtName) {
                    var evt = new MouseEvent(evtName, {
                        bubbles: true,
                        cancelable: true,
                        view: window
                    });
                    el.dispatchEvent(evt);
                });
            """, match_el)
            time.sleep(1.5)

            # Verifikasi
            val = driver.execute_script("""
                var s = document.querySelector('#WarehouseGroup .jr-mSingleselect-input-selection');
                return s ? s.textContent.trim() : '---';
            """)

            if val and val not in ('---', ''):
                print(f"  ✅  Warehouse Group terpilih: '{val}'")
                return True

            print(f"    ⚠️  Nilai belum berubah, coba lagi...")
            time.sleep(0.5)

        print(f"  ❌  Gagal memilih '{item_text}' setelah 3 percobaan")
        return False

    except Exception as e:
        print(f"  ❌  Error: {e}")
        return False

def validate_dates_v74(driver):
    driver.switch_to.default_content()
    result = driver.execute_script(r"""
        var d={sv:'',ev:''};
        var inps=document.querySelectorAll('input.date.hasDatepicker');
        if(inps[0]) d.sv=inps[0].value.trim();
        if(inps[1]) d.ev=inps[1].value.trim();
        return d;""")
    sv, ev = result['sv'], result['ev']
    so, eo = bool(sv), bool(ev)
    print(f"  🔍  Validasi: Start='{sv}' {'✅' if so else '❌'}  End='{ev}' {'✅' if eo else '❌'}")
    return so, eo

def run_cell2(driver, gc):
    print("\n" + "="*60)
    print("  🤖  CELL 2 — BOT v74 : Material Transaction Summary")
    print("="*60)
    try:
        driver.get(BOT74_REPORT_URL)
        print("  ⏳  25s tunggu load ..."); time.sleep(25)
        wait_ready(driver)
        print("\n  📋  Input Controls ...")
        fill_date_v74(driver, "Start Date", 0); time.sleep(0.8)
        fill_date_v74(driver, "End Date",   1); time.sleep(0.8)
        select_warehouse_group_v74(driver, BOT74_WAREHOUSE_GROUP)
        so, eo = validate_dates_v74(driver)
        if not so or not eo: raise SystemExit("VALIDASI TANGGAL GAGAL")
        click_apply_dialog(driver)
        wait_loading(driver)
        time.sleep(3)
        downloaded = export_xlsx(driver)
        if downloaded:
            exp  = save_to_export(downloaded, "MaterialTransactionSummary")
            url  = save_to_gsheet(gc, downloaded, "Data", "MTS")
            bot_footer(exp, url, "Data")
        else:
            print("\n  ⚠️  Download gagal")
    except SystemExit as se: print(f"\n  🛑  {se}")
    except Exception as e:   print(f"\n  ❌  {e}\n{traceback.format_exc()}")
# =============================================================================
# CELL 3 — Monitor Status Inventory Move In Progress Real Time → tab "MM IP"
# =============================================================================
BOT75IM_REPORT_URL = (
    f"{BASE_URL}/flow.html?_flowId=viewReportFlow&_flowId=viewReportFlow"
    "&ParentFolderUri=%2FiDempiere%2FLogistik%2FMonitorTrx%2FInventory_Move"
    "&reportUnit=%2FiDempiere%2FLogistik%2FMonitorTrx%2FInventory_Move%2FMonitor_Status_Inventory_Move_In_Progress__Real_Time_"
    "&standAlone=true"
)
def fill_date_dialog(driver, label, index):
    print(f"  📅  {label} → '{TODAY_STR}'")
    driver.switch_to.default_content()
    try: driver.execute_script(
        "var dp=document.querySelector('.ui-datepicker');if(dp)dp.style.display='none';")
    except: pass
    time.sleep(0.3)
    inp = None
    inps = driver.find_elements(By.CSS_SELECTOR, "input.date.hasDatepicker")
    if index < len(inps): inp = inps[index]
    if not inp:
        try:
            all_inps = driver.find_elements(By.CSS_SELECTOR,
                ".jr-mDialog input[type='text'], [class*='dialog'] input[type='text']")
            if index < len(all_inps): inp = all_inps[index]
        except: pass
    if not inp: print(f"  ❌  Input index {index} tidak ditemukan!"); return False
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp); time.sleep(0.4)
    try:
        ActionChains(driver).move_to_element(inp).click().perform(); time.sleep(0.3)
        inp.send_keys(Keys.CONTROL+"a"); time.sleep(0.1)
        inp.send_keys(Keys.DELETE);     time.sleep(0.1)
        inp.send_keys(TODAY_STR);       time.sleep(0.3)
        inp.send_keys(Keys.TAB);        time.sleep(0.5)
        val = inp.get_attribute('value')
        if val and val.strip(): trigger_events(driver, inp); print(f"  ✅  '{val}'"); return True
    except Exception as e: print(f"  ⚠️  S1: {e}")
    try:
        driver.execute_script("""
            var el=arguments[0],v=arguments[1];
            var s=Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype,'value').set;
            s.call(el,v); el.value=v;
            ['focus','input','change','blur'].forEach(function(e){
                el.dispatchEvent(new Event(e,{bubbles:true}));});
            el.dispatchEvent(new KeyboardEvent('keyup',{bubbles:true,key:'Tab',keyCode:9}));
        """, inp, TODAY_STR); time.sleep(0.5)
        val = inp.get_attribute('value')
        if val and val.strip(): print(f"  ✅  JS '{val}'"); return True
    except Exception as e: print(f"  ⚠️  S2: {e}")
    print(f"  ❌  {label} GAGAL!"); return False
def select_dropdown_by_text(driver, toggle_index, target_text):
    """Pilih dropdown ke-N (dari visible toggles) dengan teks target."""
    print(f"  🔽  Dropdown [{toggle_index}] → '{target_text}'")
    driver.switch_to.default_content()
    try:
        toggles = driver.find_elements(By.CSS_SELECTOR, "a.jr-mSingleselect-input")
        visible  = [t for t in toggles if t.is_displayed()]
        if toggle_index >= len(visible):
            print(f"  ❌  Toggle index {toggle_index} tidak ada (total: {len(visible)})")
            return False
        toggle = visible[toggle_index]
    except Exception as e:
        print(f"  ❌  {e}"); return False

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", toggle)
    time.sleep(0.5)
    tr = driver.execute_script("""
        var r=arguments[0].getBoundingClientRect();
        return {x:Math.round(r.left+r.width/2),y:Math.round(r.top+r.height/2)};
    """, toggle)
    do_click(driver, toggle, tr['x'], tr['y'])
    time.sleep(2)

    for attempt in range(3):
        items = driver.execute_script(f"""
            var res=[], target=arguments[0];
            document.querySelectorAll('a,li,div,span,option').forEach(function(el){{
                if(!el.offsetParent) return;
                var t=el.textContent.trim();
                if(t!==target) return;
                var b=el.getBoundingClientRect();
                if(b.width>=15&&b.height>=10)
                    res.push({{x:Math.round(b.x+b.width/2),y:Math.round(b.y+b.height/2)}});
            }}); return res;
        """, target_text)

        if items:
            ix, iy = items[0]['x'], items[0]['y']
            print(f"  → Klik: '{target_text}' ({ix},{iy})")
            driver.execute_script("""
                var x=arguments[0],y=arguments[1],el=document.elementFromPoint(x,y); if(!el) return;
                var o={bubbles:true,cancelable:true,view:window,clientX:x,clientY:y,
                       screenX:x,screenY:y,button:0,buttons:1};
                ['mouseover','mouseenter','mousemove','mousedown','mouseup','click']
                .forEach(function(ev){ el.dispatchEvent(new MouseEvent(ev,o)); });
            """, ix, iy)
            time.sleep(1)

            # Verifikasi nilai terpilih
            val = driver.execute_script("""
                var sels=document.querySelectorAll('.jr-mSingleselect-input-selection');
                for(var i=sels.length-1;i>=0;i--){
                    var t=sels[i].textContent.trim();
                    if(t&&t!=='---') return t;
                } return '';
            """)
            if val:
                print(f"  ✅  Terpilih: '{val}'"); return True

        if attempt < 2:
            time.sleep(1)

    print(f"  ⚠️  '{target_text}' tidak ditemukan, lanjut..."); return True
def run_cell3(driver, gc):
    print("\n" + "="*60)
    print("  🤖  BOT — Inventory Move (Pengepokan) : In Progress")
    print("="*60)
    try:
        driver.get(BOT75IM_REPORT_URL)
        print("  ⏳  25s tunggu load ..."); time.sleep(25)
        wait_ready(driver)

        print("\n  📋  Input Controls ...")

        # 1. Date Start (index 0) & Date End (index 1) — sama seperti sebelumnya
        fill_date_dialog(driver, "Date Start", 0); time.sleep(0.8)
        fill_date_dialog(driver, "Date End",   1); time.sleep(0.8)

        # 2. Branch From — dropdown index 1 (index 0 = Organization)
        select_dropdown_by_text(driver, 1, "01"); time.sleep(0.8)

        # 3. Document Type → "Inventory Move (Pengepokan)" — dropdown index 2
        select_dropdown_by_text(driver, 2, "Inventory Move (Pengepokan)"); time.sleep(0.8)

        # 4. DocStatus → "In Progress" — dropdown index 3
        select_dropdown_by_text(driver, 3, "In Progress"); time.sleep(0.8)

        # 5. Klik Apply
        click_apply_dialog(driver)
        wait_loading(driver)
        time.sleep(3)

        # 6. Export
        downloaded = export_xlsx(driver)
        if downloaded:
            exp = save_to_export(downloaded, "InventoryMove_InProgress")
            url = save_to_gsheet(gc, downloaded, "IM_IP", "Inventory Move In Progress")
            bot_footer(exp, url, "IM_IP")
        else:
            print("\n  ⚠️  Download gagal")

    except Exception as e:
        print(f"\n  ❌  {e}\n{traceback.format_exc()}")

# =============================================================================
# CELL 4 — Monitor SJ In Progress IP → tab "IP"
# =============================================================================
BOT75IP_REPORT_URL = (
    f"{BASE_URL}/flow.html?_flowId=viewReportFlow&_flowId=viewReportFlow"
    "&ParentFolderUri=%2FiDempiere%2FInventory%2FStock"
    "&reportUnit=%2FiDempiere%2FInventory%2FStock%2FMonitor_Status_Surat_Jalan___In_Progress__"
    "&standAlone=true"
)

def select_branch_ip(driver):
    print("  🏢  Branch: pilih item teratas (Jakarta)")
    driver.switch_to.default_content()
    toggle_clicked = False
    for sel in ["a.jr-mSingleselect-input", "button.jr-mSingleselect-input", ".jr-mSingleselect-input"]:
        try:
            els = driver.find_elements(By.CSS_SELECTOR, sel)
            visible = [e for e in els if e.is_displayed()]
            if visible:
                el = visible[0]
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el); time.sleep(0.4)
                r = driver.execute_script("""
                    var b=arguments[0].getBoundingClientRect();
                    return {x:Math.round(b.x+b.width/2),y:Math.round(b.y+b.height/2)};""", el)
                do_click(driver, el, r['x'], r['y']); time.sleep(2.5)
                toggle_clicked = True; break
        except: continue
    if not toggle_clicked:
        try:
            el = driver.execute_script("return document.elementFromPoint(914,295);")
            if el: do_click(driver, el, 914, 295); time.sleep(2.5)
        except: pass
    for attempt in range(3):
        items = driver.execute_script(r"""
            var res=[], SKIP=['---','','–','-','*'];
            var containers=document.querySelectorAll(
                '.jr-mSingleselect-listbox,.jr-mSingleselect-dropdown,[class*="listbox"],[class*="dropDown"]');
            containers.forEach(function(c){
                if(!c.offsetParent) return;
                c.querySelectorAll('li,a,div[role="option"],span').forEach(function(el){
                    if(!el.offsetParent) return;
                    var t=el.textContent.trim();
                    if(!t||SKIP.indexOf(t)>-1||t.length>80) return;
                    if(/^[\*\-–\s]+$/.test(t)) return;
                    var b=el.getBoundingClientRect();
                    if(b.width<5||b.height<5) return;
                    res.push({text:t,cx:Math.round(b.left+b.width/2),cy:Math.round(b.top+b.height/2)});
                });
            });
            if(res.length===0){
                document.querySelectorAll('li').forEach(function(el){
                    if(!el.offsetParent) return;
                    var t=el.textContent.trim();
                    if(!t||SKIP.indexOf(t)>-1||t.length>80) return;
                    if(/^[\*\-–\s]+$/.test(t)) return;
                    var b=el.getBoundingClientRect();
                    if(b.width<5||b.height<5||b.left<480||b.left>970||b.top<190||b.top>710) return;
                    res.push({text:t,cx:Math.round(b.left+b.width/2),cy:Math.round(b.top+b.height/2)});
                });
            }
            return res;""")
        if items:
            first = items[0]; ix, iy = first['cx'], first['cy']
            print(f"  → Klik: '{first['text']}' ({ix},{iy})")
            driver.execute_script("""
                var x=arguments[0],y=arguments[1],el=document.elementFromPoint(x,y); if(!el) return;
                var o={bubbles:true,cancelable:true,view:window,clientX:x,clientY:y,screenX:x,screenY:y,button:0,buttons:0};
                el.dispatchEvent(new MouseEvent('mouseover',o));
                el.dispatchEvent(new MouseEvent('mouseenter',o));
                el.dispatchEvent(new MouseEvent('mousemove',o));""", ix, iy); time.sleep(0.5)
            driver.execute_script("""
                var x=arguments[0],y=arguments[1],el=document.elementFromPoint(x,y); if(!el) return;
                var o={bubbles:true,cancelable:true,view:window,clientX:x,clientY:y,screenX:x,screenY:y,button:0,buttons:1};
                el.dispatchEvent(new MouseEvent('mousedown',o));
                el.dispatchEvent(new MouseEvent('mouseup',o));
                el.dispatchEvent(new MouseEvent('click',o));""", ix, iy); time.sleep(1.2)
            closed = driver.execute_script(r"""
                var d=document.querySelectorAll('.jr-mSingleselect-listbox,.jr-mSingleselect-dropdown,[class*="listbox"],[class*="dropDown"]');
                for(var i=0;i<d.length;i++) if(d[i].offsetParent) return false;
                return true;""")
            if closed: print(f"  ✅  Branch '{first['text']}' dipilih"); return True
            val = driver.execute_script(r"""
                var s=document.querySelectorAll('.jr-mSingleselect-input-selection,[class*="singleSelect"] [class*="value"]');
                for(var i=0;i<s.length;i++){var t=s[i].textContent.trim(); if(t&&t!=='---') return t;}
                return '';""")
            if val and val != '---': print(f"  ✅  Branch: '{val}'"); return True
        if attempt < 2: time.sleep(1.5)
    print("  ⚠️  Branch: tidak bisa konfirmasi, lanjut ..."); return True

def run_cell4(driver, gc):
    print("\n" + "="*60)
    print("  🤖  CELL 4 — BOT v75 IP : Monitor SJ In Progress (IP)")
    print("="*60)
    try:
        driver.get(BOT75IP_REPORT_URL)
        print("  ⏳  25s tunggu load ..."); time.sleep(25)
        wait_ready(driver)
        print("\n  📋  Input Controls ...")
        select_branch_ip(driver); time.sleep(0.8)
        click_apply_dialog(driver)
        wait_loading(driver)
        time.sleep(3)
        downloaded = export_xlsx(driver)
        if downloaded:
            exp = save_to_export(downloaded, "MonitorSuratJalan_IP")
            url = save_to_gsheet(gc, downloaded, "IP", "Monitor SJ IP")
            bot_footer(exp, url, "IP")
        else:
            print("\n  ⚠️  Download gagal")
    except Exception as e: print(f"\n  ❌  {e}\n{traceback.format_exc()}")

# =============================================================================
# CELL 5 — iDempiere ERP → tab "IP_iDempiere"
# =============================================================================
def convert_xls_to_xlsx(xls_path):
    xlsx_path = xls_path.replace(".xls", ".xlsx")
    if xlsx_path == xls_path: return xls_path
    print(f"   🔄 Konversi {os.path.basename(xls_path)} → .xlsx ...")
    wb_xls  = xlrd.open_workbook(xls_path)
    wb_xlsx = openpyxl.Workbook()
    for sheet_idx in range(wb_xls.nsheets):
        ws_xls  = wb_xls.sheet_by_index(sheet_idx)
        ws_xlsx = wb_xlsx.active if sheet_idx == 0 else wb_xlsx.create_sheet()
        ws_xlsx.title = ws_xls.name
        for row in range(ws_xls.nrows):
            for col in range(ws_xls.ncols):
                ws_xlsx.cell(row=row+1, column=col+1, value=ws_xls.cell_value(row, col))
    wb_xlsx.save(xlsx_path)
    print(f"   ✅ Konversi selesai → {os.path.basename(xlsx_path)}")
    return xlsx_path

def fill_text_field_erp(driver, c5_wait, label_text, text_value):
    print(f"   → Mengisi {label_text}: {text_value}")
    xpath = (f"//span[contains(text(), '{label_text}')]"
             f"/ancestor::tr[1]//input[not(@type='hidden')]")
    for attempt in range(3):
        try:
            inputs = driver.find_elements(By.XPATH, xpath)
            if not inputs:
                inputs = driver.find_elements(By.XPATH,
                    f"//span[contains(text(), '{label_text}')]"
                    f"/following::input[not(@type='hidden')]")
            if not inputs: print(f"   ⚠️ Gagal: {label_text} tidak ditemukan."); return
            inp = inputs[0]
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", inp)
            time.sleep(0.5)
            inp.click(); inp.clear(); time.sleep(0.5)
            for char in text_value:
                inp.send_keys(char); time.sleep(0.15)
            time.sleep(3); inp.send_keys(Keys.ENTER); time.sleep(2); return
        except Exception:
            print(f"   🔄 Retry {label_text} ({attempt+1})..."); time.sleep(2)

def select_date_erp(driver, label_text, index):
    print(f"   → Memilih {label_text} (kotak ke-{index}): Hari Ini")
    xpath = (
        f"//span[contains(text(), '{label_text}')]"
        f"/ancestor::tr[1]//i[contains(@class, 'z-icon-calendar') "
        f"or contains(@class, 'datebox-icon')] | "
        f"//span[contains(text(), '{label_text}')]"
        f"/ancestor::tr[1]//a[contains(@class, 'datebox-button')]"
    )
    try:
        icons = driver.find_elements(By.XPATH, xpath)
        if not icons:
            icons = driver.find_elements(By.XPATH,
                f"//span[contains(text(), '{label_text}')]"
                f"/following::i[contains(@class, 'z-icon-calendar') "
                f"or contains(@class, 'datebox-icon')]")
        target_icon = icons[index - 1]
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", target_icon)
        time.sleep(0.5); driver.execute_script("arguments[0].click();", target_icon); time.sleep(1.5)
        today_day_str = str(datetime.now().day)
        xpath_tgl = (
            f"//div[contains(@class, 'calendar') or contains(@class, 'popup')]"
            f"//td[normalize-space(text())='{today_day_str}' "
            f"and not(contains(@class, 'disd')) "
            f"and not(contains(@class, 'outside'))]"
        )
        for cell in driver.find_elements(By.XPATH, xpath_tgl):
            if cell.is_displayed():
                driver.execute_script("arguments[0].click();", cell); break
        time.sleep(1.5)
    except Exception as e:
        print(f"   ⚠️ Gagal memilih tanggal {label_text}: {e}")

def run_cell5(driver, gc):
    print("\n" + "="*60)
    print("  🤖  CELL 5 — iDempiere ERP → tab 'IP_iDempiere'")
    print("="*60)

    # Alihkan CDP ke folder ERP
    try:
        driver.execute_cdp_cmd("Page.setDownloadBehavior",
            {"behavior": "allow", "downloadPath": C5_DOWNLOAD_DIR})
        print(f"  ✅  CDP dialihkan → {C5_DOWNLOAD_DIR}")
    except Exception as e:
        print(f"  ⚠️  CDP: {e}")

    c5_wait = WebDriverWait(driver, 20)

    try:
        # 1. LOGIN ERP
        print("  🌐  Membuka halaman login iDempiere ...")
        driver.get(ERP_URL); time.sleep(5)
        print("  🔑  Proses Login ...")
        user_xpath = "(//input[contains(@class, 'z-textbox') or @type='text'])[1]"
        pass_xpath = "//input[@type='password']"
        user_input = c5_wait.until(EC.element_to_be_clickable((By.XPATH, user_xpath)))
        pass_input = c5_wait.until(EC.element_to_be_clickable((By.XPATH, pass_xpath)))
        user_input.clear(); user_input.send_keys(ERP_USER)
        pass_input.clear(); pass_input.send_keys(ERP_PASS)
        login_btn = driver.find_element(By.XPATH,
            "//button[contains(., 'OK')] | //div[contains(@class, 'login-btn')]")
        driver.execute_script("arguments[0].click();", login_btn)
        print("  ⏳  Menunggu workspace (15s) ..."); time.sleep(15)

        # 2. BUKA MENU TRANSACTION DETAIL
        print("  📂  Membuka menu Transaction Detail ...")
        menu_item = c5_wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//span[normalize-space(text())='Transaction Detail']")))
        driver.execute_script("arguments[0].click();", menu_item); time.sleep(8)

        # 3. PILIH TANGGAL
        select_date_erp(driver, "Movement Date", 1)
        select_date_erp(driver, "Movement Date", 2)

        # 4. KLIK OK
        print("  🚀  Klik tombol OK ...")
        ok_xpath = (
            "//button[contains(translate(normalize-space(.), 'ok', 'OK'), 'OK') "
            "or @title='OK'] | //a[contains(translate(normalize-space(.), 'ok', 'OK'), 'OK')]"
        )
        clicked = False
        for btn in reversed(driver.find_elements(By.XPATH, ok_xpath)):
            if btn.is_displayed():
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                time.sleep(1); driver.execute_script("arguments[0].click();", btn)
                clicked = True; print("   ✅ OK diklik!"); break
        if not clicked: raise Exception("Tombol OK tidak ditemukan.")
        print("  ⏳  Generate report (20s) ..."); time.sleep(20)

        # 6. FORMAT XLS & DOWNLOAD
        print("  🔄  Format → XLS ...")
        pdf_selects = driver.find_elements(By.XPATH, "//select[option[contains(text(), 'PDF')]]")
        if pdf_selects:
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", pdf_selects[0])
            time.sleep(1); Select(pdf_selects[0]).select_by_visible_text("XLS")
        else:
            pdf_icon_xpath = (
                "//input[@value='PDF']/following-sibling::* | "
                "//span[text()='PDF']/ancestor::*[contains(@class, 'combo')]//i"
            )
            pdf_btn = c5_wait.until(EC.element_to_be_clickable((By.XPATH, pdf_icon_xpath)))
            driver.execute_script("arguments[0].click();", pdf_btn); time.sleep(1.5)
            xls_opt = c5_wait.until(EC.element_to_be_clickable((By.XPATH,
                "//li[text()='XLS'] | //span[text()='XLS'] | "
                "//div[contains(@class, 'comboitem') and text()='XLS']")))
            driver.execute_script("arguments[0].click();", xls_opt)
        print("  ✅  Format → XLS"); time.sleep(3)

        print("  📥  Klik 'Save to File' ...")
        save_btn_xpath = (
            "//*[contains(text(), 'Save to File')]/ancestor-or-self::button | "
            "//button[contains(., 'Save to File')]"
        )
        save_btns = driver.find_elements(By.XPATH, save_btn_xpath)
        if save_btns:
            driver.execute_script("arguments[0].click();", save_btns[0])
        else:
            driver.find_element(By.XPATH, "//*[text()='Save to File']").click()
        print("  ✅  Save to File diklik"); time.sleep(10)

    except Exception as e:
        print(f"\n  ❌  {e}\n{traceback.format_exc()}")

    # 7. KONVERSI + UPLOAD (di luar try agar selalu dijalankan)
    try:
        files = os.listdir(C5_DOWNLOAD_DIR)
        if files:
            full_paths  = [os.path.join(C5_DOWNLOAD_DIR, f) for f in files]
            latest_file = max(full_paths, key=os.path.getmtime)
            print(f"\n  ✅  File ditemukan: {os.path.basename(latest_file)}")
            if latest_file.lower().endswith(".xls") and not latest_file.lower().endswith(".xlsx"):
                latest_file = convert_xls_to_xlsx(latest_file)
            exp = save_to_export(latest_file, "MonitorSuratJalan_IP_iDempiere")
            url = save_to_gsheet(gc, latest_file, "IP_iDempiere", "Monitor SJ IP (iDempiere)")
            bot_footer(exp, url, "IP_iDempiere")
        else:
            print("\n  ⚠️  Tidak ada file di folder download ERP.")
    except Exception as e:
        print(f"\n  ⚠️  Gagal konversi/upload: {e}")

# =============================================================================
# MAIN — 1 driver, 1 login, 4 tab
# =============================================================================
def run_all_shared(gc, cells):
    driver = make_driver(DOWNLOAD_DIR)
    try:
        # Set CDP download dir awal
        try:
            driver.execute_cdp_cmd("Page.setDownloadBehavior",
                {"behavior": "allow", "downloadPath": DOWNLOAD_DIR})
            print(f"  ✅  CDP download path: {DOWNLOAD_DIR}")
        except Exception as e:
            print(f"  ⚠️  CDP: {e}")

        # LOGIN 1x untuk cell 2/3/4 (Jasper)
        jasper_cells = [c for c in cells if c in (2, 3, 4)]
        erp_cells    = [c for c in cells if c == 5]

        if jasper_cells:
            print("\n" + "="*60)
            print("  🔑  LOGIN JASPER (1x untuk semua Cell 2/3/4)")
            print("="*60)
            do_login(driver)

        first_tab = driver.window_handles[0]

        for cell in cells:
            if cell in (2, 3, 4):
                # Pastikan CDP kembali ke jasper downloads
                try:
                    driver.execute_cdp_cmd("Page.setDownloadBehavior",
                        {"behavior": "allow", "downloadPath": DOWNLOAD_DIR})
                except: pass
                open_new_tab(driver)
                if   cell == 2: run_cell2(driver, gc)
                elif cell == 3: run_cell3(driver, gc)
                elif cell == 4: run_cell4(driver, gc)
                driver.switch_to.window(first_tab)
                print(f"  ↩️   Kembali ke tab utama")

            elif cell == 5:
                open_new_tab(driver)
                run_cell5(driver, gc)   # CDP dialihkan di dalam run_cell5
                driver.switch_to.window(first_tab)
                print(f"  ↩️   Kembali ke tab utama")

    finally:
        try: driver.quit()
        except: pass
        print("\n  🔒  Browser ditutup (1x)")


if __name__ == "__main__":
    print(f"🤖 JasperBot START — {datetime.now(WIB).strftime('%Y-%m-%d %H:%M:%S WIB')}")
    cells = [2, 3, 4, 5]
    print(f"   Cell    : {cells}")
    print(f"   Mode    : 1 browser → 1 login → {len(cells)} tab")

    gc = init_gc()
    run_all_shared(gc, cells)

    print(f"\n🏁 JasperBot SELESAI — {datetime.now(WIB).strftime('%Y-%m-%d %H:%M:%S WIB')}")
